import torch
import torch.utils.data as data
import torchvision
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import csv
import random
from openpyxl import load_workbook
import cv2
from torchvision import transforms

class Kadid10kFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num, patch_size=224, self_patch_num=1, use_L=False):
        self.patch_size = patch_size
        self.self_patch_num = self_patch_num
        self.use_L = use_L
        self.transform = transform

        imgname = []
        refimgname = []
        mos_all = []
        csv_file = os.path.join(root, 'dmos.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['dist_img'])
                refimgname.append(row['ref_img'])
                mos = np.array(float(row['dmos'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'images', imgname[item]),os.path.join(root, 'images', refimgname[item]), mos_all[item]))

        self.samples = sample


    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]

        LQ = pil_loader(LQ_path, self.use_L)
        LQ_patches = []
        for _ in range(self.self_patch_num):
            LQ_patch = self.transform(LQ)
            LQ_patches.append(LQ_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)

        return LQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length

class LIVEFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num, patch_size=224, self_patch_num=1, use_L=False):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num
        self.transform = transform
        self.use_L = use_L

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = scipy.io.loadmat(os.path.join(root, 'dmos_realigned.mat'))
        labels = dmos['dmos_new'].astype(np.float32)

        orgs = dmos['orgs']
        refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refnames_all = refnames_all['refnames_all']

        sample = []
        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    LQ_path = imgpath[item]
                    HQ_path = os.path.join(root, 'refimgs', refnames_all[0][item][0])
                    label = labels[0][item]
                    sample.append((LQ_path, HQ_path, label))

        self.samples = sample


    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path, self.use_L)
        LQ_patches = []
        for _ in range(self.self_patch_num):
            LQ_patch = self.transform(LQ)
            LQ_patches.append(LQ_patch.unsqueeze(0))
        LQ_patches = torch.cat(LQ_patches, 0)

        return LQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = '{:0>3d}{}'.format(index, '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename

class CSIQFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num, patch_size =224, self_patch_num=1, use_L=False):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num
        self.transform = transform
        self.use_L = use_L

        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath,'.png')
        txtpath = os.path.join(root, 'csiq_label.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[0]))
            target.append(words[1])
            ref_temp = words[0].split(".")
            refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []

        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    LQ_path = os.path.join(root, 'dst_imgs_all', imgnames[item])
                    HQ_path = os.path.join(root, 'src_imgs', refnames_all[item])
                    label = labels[item]
                    sample.append((LQ_path, HQ_path, label))
        self.samples = sample


    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path, self.use_L)
        LQ_patches = []
        for _ in range(self.self_patch_num):
            LQ_patch = self.transform(LQ)
            LQ_patches.append(LQ_patch.unsqueeze(0))
        LQ_patches = torch.cat(LQ_patches, 0)

        return LQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length

class TID2013Folder(data.Dataset):

    def __init__(self, root, index, transform, patch_num, patch_size=224, self_patch_num=1, use_L=False):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num
        self.transform = transform
        self.use_L = use_L
        
        refpath = os.path.join(root, 'reference_images')
        refname = self._getTIDFileName(refpath,'.bmp.BMP')
        txtpath = os.path.join(root, 'mos_with_names.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refnames_all.append(ref_temp[0][1:])
        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    LQ_path = os.path.join(root, 'distorted_images', imgnames[item])
                    HQ_name = 'I' + imgnames[item].split("_")[0][1:] + '.BMP'
                    HQ_path = os.path.join(refpath, HQ_name)
                    label = labels[item]
                    sample.append((LQ_path, HQ_path, label))
        self.samples = sample


    def _getTIDFileName(self, path, suffix):
        filename = []
        f_list = os.listdir(path)
        for i in f_list:
            if suffix.find(os.path.splitext(i)[1]) != -1:
                filename.append(i[1:3])
        return filename

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, HQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, HQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path, self.use_L)
        LQ_patches = []
        for _ in range(self.self_patch_num):
            LQ_patch = self.transform(LQ)
            LQ_patches.append(LQ_patch.unsqueeze(0))
        LQ_patches = torch.cat(LQ_patches, 0)

        return LQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length

class LIVEChallengeFolder(data.Dataset):
    def __init__(self, root, index, transform, patch_num, patch_size=224, self_patch_num=1, use_L=False):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num
        self.transform = transform
        self.use_L = use_L

        LQ_pathes = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        LQ_pathes = LQ_pathes['AllImages_release']
        LQ_pathes = LQ_pathes[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, 'Images', LQ_pathes[item][0][0]), labels[item]))
        self.samples = sample
    
    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path, self.use_L)
        LQ_patches = []
        for _ in range(self.self_patch_num):
            LQ_patch = self.transform(LQ)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)

        return LQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length

class BIDChallengeFolder(data.Dataset):
    def __init__(self, root, index, transform, patch_num, patch_size=224, self_patch_num=1, use_L=False):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num
        self.transform = transform
        self.use_L = use_L

        LQ_pathes = []
        labels = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for _ in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            LQ_pathes.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            labels.append(mos)
            if count == 587:
                break

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, LQ_pathes[item]), labels[item]))
        self.samples = sample
    
    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path, self.use_L)
        LQ_patches = []
        for _ in range(self.self_patch_num):
            LQ_patch = self.transform(LQ)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)

        return LQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length
    
class Koniq_10kFolder(data.Dataset):
    def __init__(self, root, index, transform, patch_num, patch_size=224, self_patch_num=1, use_L=False):
        self.patch_size =patch_size
        self.self_patch_num = self_patch_num
        self.transform = transform
        self.use_L = use_L

        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_scores_and_distributions.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS_zscore'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append((os.path.join(root, '1024x768', imgname[item]), mos_all[item]))
        
        self.samples = sample

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (LQ, target) where target is IQA values of the target LQ.
        """
        LQ_path, target = self.samples[index]
        LQ = pil_loader(LQ_path, self.use_L)
        LQ_patches = []
        for _ in range(self.self_patch_num):
            LQ_patch = self.transform(LQ)
            
            LQ_patches.append(LQ_patch.unsqueeze(0))
        #[self_patch_num, 3, patch_size, patch_size]
        LQ_patches = torch.cat(LQ_patches, 0)

        return LQ_patches, target

    def __len__(self):
        length = len(self.samples)
        return length


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getPairRandomPatch(img1, img2, crop_size=512):
    (iw,ih) = img1.size
    # print(ih,iw)

    ip = int(crop_size)

    ix = random.randrange(0, iw - ip + 1)
    iy = random.randrange(0, ih - ip + 1)


    img1_patch = img1.crop((ix, iy, ix+ip, iy+ip))#左上右下
    img2_patch = img2.crop((ix, iy, ix+ip, iy+ip))#左上右下

    return img1_patch, img2_patch

def getPairAugment(img1, img2, hflip=True, vflip=True, rot=True):
    hflip = hflip and random.random() < 0.5
    vflip = vflip and random.random() < 0.5
    rot180 = rot and random.random() < 0.5

    if hflip: 
        img1 = img1.transpose(Image.FLIP_TOP_BOTTOM)
        img2 = img2.transpose(Image.FLIP_TOP_BOTTOM)
    if vflip: 
        img1 = img1.transpose(Image.FLIP_LEFT_RIGHT)
        img2 = img2.transpose(Image.FLIP_LEFT_RIGHT)
    if rot180: 
        img1 = img1.transpose(Image.ROTATE_180)
        img2 = img2.transpose(Image.ROTATE_180)
        
    return img1, img2


def getSelfPatch(img, patch_size, patch_num, is_random=True):
    (iw,ih) = img.size
    patches = []
    for i in range(patch_num):
        if is_random:
            ix = random.randrange(0, iw - patch_size + 1)
            iy = random.randrange(0, ih - patch_size + 1)
        else:ix,iy=(iw - patch_size + 1)//2,(ih - patch_size + 1)//2

        # patch = img[iy:iy + lr_size, ix:ix + lr_size, :]#上下左右
        patch = img.crop((ix, iy, ix+patch_size, iy+patch_size))#左上右下
        patches.append(patch)

    return patches


def pil_loader(path, use_L=False):
    if use_L:
        return Image.open(path).convert('L')
    else:
        return Image.open(path).convert('RGB')
